import json
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from utilities import device_filter, variable_filter
from constants import *
from user_settings import *

# import json file into python
with open(json_input_file, 'r') as MetaVariables_JSON:
    jsondata = json.load(MetaVariables_JSON)

# Excel file manipulation
# # open excel file and select sheet
workbook = load_workbook(file_path, data_only=True)  # load file to excel
if active_sheet is None:
    worksheet = workbook.active
else:
    try:
        worksheet = workbook[active_sheet]
    except KeyError:
        # sheet doesn't exist, so create it
        workbook.create_sheet(active_sheet)
        worksheet = workbook[active_sheet]
    except:
        print("Error: could not open/create worksheet")
        quit(3)

# check that the user has chosen the correct files
print(f"Excel MetaVariable file: {cblu}{file_path}{cblk}")
if active_sheet is not None: print(f"Excel MetaVariable sheet: {clbl}{active_sheet}{cblk}")
print(f"JSON MetaVariable file: {cgrn}{json_input_file}{cblk}")
print(f"Device column header: {cylw}{HOSTNAME}{cblk}  |  VDOM column header: {corg}{VDOM}{cblk}")
user_choice = input("Please confirm above settings: [y/n]\n")
if user_choice not in ['y', 'Y']:
    print("SCRIPT ABORTED")
    quit(9)

# compare ADOM in JSON file to ADOM in user_settings.py
if jsondata['adom'] != adom:
    print(f"ADOM from JSON file '{cgrn}{jsondata['adom']}{cblk}' does not match adom in user settings '{cblu}{adom}{cblk}'")
    user_choice = print("Are you sure you want to continue? [y/n]")
    print('\n')
    if user_choice not in ['y', 'Y']:
        print("SCRIPT ABORTED")
        quit(7)

# create a backup of the initial excel file
isbackupfail = False  # boolean for fail contingency
backup_file, file_type = file_path.rsplit('.', 1)
backup_file = f"{backup_file}-backup.{file_type}"  # name backup file
try:
    shutil.copy2(file_path, backup_file)  # create backup file
except PermissionError:
    print("ERROR: permission denied.")  # call out error if there is no permission to copy file
    isbackupfail = True  # enable boolean for fail contingency
except:
    print("ERROR: could not create backup file.")  # call out general copy error
    isbackupfail = True  # enable boolean for fail contingency
if isbackupfail:
    user_choice = input(
        "Could not create backup of Excel MetaVariable file.  Are you sure you want to continue? [y/n]")  # get user input to continue
    print('\n')
    if user_choice not in ['y', 'Y']:
        print("SCRIPT ABORTED")
        quit(8)

# In Excel File, set host/vdom columns and check for duplicate entries
duplicate_check = []
host_column = None
vdom_column = None
for column in worksheet.iter_cols():
    # assign friendly names
    header = column[0].value
    current_column = column[0].column
    # check for hostname column
    if header == HOSTNAME:
        host_column = current_column
    # check for vdom column
    if header == VDOM:
        vdom_column = current_column
    # add value to duplicate_check list
    if header not in skipable:
        if header in duplicate_check:
            print(f"Error: variable \"{header}\" is duplicated at column {get_column_letter(current_column)}.")
            quit(2)
        duplicate_check.append(header)
# # # if host column is not present, create it
if host_column is None:
    print(f"No {HOSTNAME} column present.  Creating it.")
    host_column = 1
    worksheet.insert_cols(host_column)
    worksheet.cell(row=1, column=host_column).value = HOSTNAME
# # # if vdom column is not present, create it
if vdom_column is None:
    print(f"No {VDOM} column present.  Creating it.")
    vdom_column = host_column + 1
    worksheet.insert_cols(vdom_column)
    worksheet.cell(row=1, column=vdom_column).value = VDOM
start_column = max([host_column, vdom_column]) + 1

# Create a fill object pattern
new_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow background color
default_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')  # Orange, accent 2

# parse and write to Excel file
for vardata in jsondata['variables']:
    # apply filter to variable
    if not variable_filter(vardata[NAME]):  # if variable does not pass the filter, skip it
        continue
    # find variable column, if it exists
    varcolumn = None
    for column in worksheet.iter_cols(min_col=start_column):
        # assign friendly names
        header = column[0].value
        current_column = column[0].column
        # check if header contains variable
        if header == vardata[NAME]:
            varcolumn = current_column
            break  # stop searching since variable column has been found
    # if variable column does not exist...
    if varcolumn is None:
        # if "add new variable" boolean is true, create new variable column
        if add_new_vars:
            varcolumn = start_column
            worksheet.insert_cols(varcolumn)
            worksheet.cell(row=1, column=varcolumn).value = vardata[NAME]
            worksheet.cell(row=1, column=varcolumn).fill = new_fill
        # if "add new variable" boolean is false, continue to next variable
        else:
            continue
    # write default value for variable column
    if GL_VALUE in vardata and vardata[GL_VALUE] not in skipable:
        worksheet.cell(row=2, column=varcolumn).value = vardata[GL_VALUE]
        worksheet.cell(row=2, column=varcolumn).fill = default_fill
    # write device specific values for variable column
    maprow = None
    if MAPPING in vardata:  # if there is no "mapping" section for the variable, this section is moot
        for mapval in vardata[MAPPING]:
            # apply filter to device
            if not device_filter(mapval[DEVICE]):  # if device does not pass the filter, skip it
                continue
            # search hostname column for device name, if it exists
            for row in worksheet.iter_rows(min_row=3):
                # assign friendly names
                header = row[host_column-1].value
                current_row = row[host_column-1].row
                # check if header contains variable
                if header == mapval[DEVICE]:
                    maprow = current_row
                    break  # stop searching since device row has been found
            # if device is not found in hostname column...
            if maprow is None:
                # if "add new devices" boolean is true, add the device
                if add_new_devices:
                    maprow = 3
                    worksheet.insert_rows(maprow)
                    worksheet.cell(row=maprow, column=host_column).value = mapval[DEVICE]
                    worksheet.cell(row=maprow, column=host_column).fill = new_fill
                # if "add new devices" boolean is false, continue to the next device
                else:
                    continue
            # write mapped data to cell
            if VIRTUAL_DOMAIN in mapval and mapval[VIRTUAL_DOMAIN] not in skipable:
                worksheet.cell(row=maprow, column=vdom_column).value = mapval[VIRTUAL_DOMAIN]
            worksheet.cell(row=maprow, column=varcolumn).value = mapval[VALUE]

workbook.save(file_path)
